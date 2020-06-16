"""update_produce.py
스프레드시트 업데이트 하기
"""
import openpyxl

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]

# The produce types and their updated prices
PRICE_UPDATES = {
    "Garlic": 3.07,
    "Celery": 1.19,
    "Lemon": 1.27,
}

# Loop through the rows and update the prices
for row_num in range(2, sheet.max_row):
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save("updated_produce_sales.xlsx")
