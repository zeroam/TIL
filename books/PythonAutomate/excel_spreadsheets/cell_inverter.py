"""cell_inverter.py
스프레드시트 셀 반전시키기
"""
import sys
import openpyxl

if len(sys.argv) != 2:
    print(f"python {__file__} <*.xlsx>")
    sys.exit(1)

file_name = sys.argv[1]

wb = openpyxl.load_workbook(file_name)
sheet = wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for row_obj in sheet.rows:
    for cell in row_obj:
        new_sheet.cell(row=cell.column, column=cell.row).value = cell.value

new_wb.save(f"reversed_{file_name}")
