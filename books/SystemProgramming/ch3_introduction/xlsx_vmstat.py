#!/bin/env python
import openpyxl
from openpyxl.styles import Alignment, Font
from subprocess import Popen, PIPE


def create_header(sheet, start_col: str, end_col: str, word: str) -> None:
    sheet.merge_cells(f"{start_col}:{end_col}")
    sheet[start_col] = word
    sheet[start_col].font = Font(bold=True)
    sheet[start_col].alignment = Alignment(horizontal="center")


def main():
    cmd = "vmstat 1 5 | awk '{now=strftime(\"%Y-%m-%d %T \"); print now $0}'"
    p = Popen(cmd, shell=True, stdout=PIPE)
    (ret, err) = p.communicate()

    wb = openpyxl.Workbook()
    sheet = wb.active
    rows = ret.decode().split("\n")

    for row_idx, row in enumerate(rows):
        # 첫 번째 해더는 추가 X
        if row_idx == 0:
            continue

        columns = row.split()
        for col_index, col in enumerate(columns):
            sheet.cell(row_idx + 1, col_index + 1).value = col

    # 헤더 추가
    create_header(sheet, "A1", "B1", "datetime")
    create_header(sheet, "C1", "D1", "procs")
    create_header(sheet, "E1", "H1", "memory")
    create_header(sheet, "I1", "J1", "swap")
    create_header(sheet, "K1", "L1", "io")
    create_header(sheet, "M1", "N1", "system")
    create_header(sheet, "O1", "S1", "cpu")

    file_name = "vmstat.xlsx"
    wb.save(file_name)
    print(f"{file_name} 파일에 저장되었습니다.")


if __name__ == "__main__":
    main()
