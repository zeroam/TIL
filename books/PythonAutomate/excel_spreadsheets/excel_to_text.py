"""excel_to_text.py
엑셀 파일을 다수의 텍스트 파일로 변환
ex) python excel_to_text.py <엑셀파일>
"""
import sys
import openpyxl
from openpyxl.utils import get_column_letter

if len(sys.argv) != 2:
    print(f"python {__file__} <excel file>")
    sys.exit(1)

file_name = sys.argv[1]
base, _, ext = file_name.rpartition(".")
wb = openpyxl.load_workbook(file_name)
sheet = wb.active

# 열 기준으로 목록 조회
for i, col_obj in enumerate(sheet.columns):
    col_name = get_column_letter(i + 1)
    text_file = f"{base}_{col_name}.txt"
    with open(text_file, "w") as f:
        for cell in col_obj:
            f.write(f"{cell.value}\n")
