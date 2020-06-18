"""multiplication_table.py
명령행에서 수 N을 받아 NxN 곱셈표를 만드는 프로그램
ex) python multiplication_table.py 6
"""
import sys
import openpyxl
from openpyxl.styles import Font

if len(sys.argv) != 2:
    print(f"python {__file__} <num>")
    sys.exit(1)

size = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
table_font = Font(bold=True, italic=True)
for i in range(1, size + 1):
    # 곱셈표 값 설정
    sheet.cell(row=1, column=i + 1).font = table_font
    sheet.cell(row=1, column=i + 1).value = i
    sheet.cell(row=i + 1, column=1).font = table_font
    sheet.cell(row=i + 1, column=1).value = i
    for j in range(1, size + 1):
        sheet.cell(row=i + 1, column=j + 1).value = i * j

wb.save("multiple.xlsx")
