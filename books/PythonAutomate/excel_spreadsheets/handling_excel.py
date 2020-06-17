"""handling_excel.py
openpyxl을 이용해 excel 파일 다루기
"""
import os
import openpyxl


# Excel 문서 열기
wb = openpyxl.load_workbook("example.xlsx")
print(f"type: {type(wb)}")  # <class 'openpyxl.workbook.workbook.Workbook'>


# 특정 Sheet 열기
print(f"Sheet Names: {wb.sheetnames}")  # ['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb["Sheet3"]
print(f"Sheet Type: {type(sheet)}")  # <class 'openpyxl.worksheet.worksheet.Worksheet'>
print(f"Sheet Title: {sheet.title}")  # Sheet3


# 활성화 된 Sheet 열기
active_sheet = wb.active
print(f"Active Sheet: {active_sheet}")  # <Worksheet "Sheet1">


# Sheet로부터 Cell 값 가져오기
a1 = active_sheet["A1"]
print(f"Cell: {a1}")  # <Cell 'Sheet1'.A1>
print(f"Cell Value: {a1.value}")
print(f"Cell (row: {a1.row}, column: {a1.column}, value: {a1.value}")
b1 = active_sheet["B1"]
print(f"Cell {b1.coordinate} is {b1.value}")

# cell 메서드로 Cell 값 가져오기
for i in range(1, 8, 2):
    print(i, active_sheet.cell(row=i, column=2).value)

# max_row, max_column 값 얻기
print(
    f"{active_sheet.title} - "
    f"max_row: {active_sheet.max_row}, max_column: {active_sheet.max_column}"
)


# Column 문자와 숫자 변환하기
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1), end=" ") # A
print(get_column_letter(2), end=" ") # B
print(get_column_letter(900), end=" ")   # AHP
print(get_column_letter(active_sheet.max_column), end=" ")   # C
print(column_index_from_string("A"), end=" ")    # 1
print(column_index_from_string("AA"))    # 27


# Cell 목록 얻어오기
print("--- START OF ROW ---")
for row_cell_obj in active_sheet["A1": "C3"]:
    for cell_obj in row_cell_obj:
        print(cell_obj.coordinate, cell_obj.value)
    print("--- END OF ROW ---")

# Cell 목록 얻어오기 - 열 기준
print("--- START OF COLUMN ---")
for cell_obj in list(active_sheet.columns)[1]:
    print(cell_obj.coordinate, cell_obj.value)
print("--- END OF COLUMN ---")

# Cell 목록 얻어오기 - 행 기준
print("--- START OF ROW ---")
for cell_obj in list(active_sheet.rows)[1]:
    print(cell_obj.coordinate, cell_obj.value)
print("--- END OF ROW ---")


# Excel 파일 생성하기
wb = openpyxl.Workbook()
print(wb.sheetnames)    # ['Sheet']
sheet = wb.active
sheet.title = "Spam Bacon Eggs Sheet"   # Change title
wb.save("sample.xlsx")
print(wb.sheetnames)    # ['Spam Bacon Eggs Sheet']


# Sheet 생성, 제거 하기
wb.create_sheet()
print(wb.sheetnames)    # ['Spam Bacon Eggs Sheet', 'Sheet']
wb.create_sheet(index=0, title="First Sheet")
print(wb.sheetnames)    # ['First Sheet', 'Spam Bacon Eggs Sheet', 'Sheet']
del wb['Spam Bacon Eggs Sheet']
del wb['First Sheet']
print(wb.sheetnames)    # ['Sheet']


# 폰트 설정하기
from openpyxl.styles import Font
sheet = wb.active
font_obj1 = Font(name="Times New Roman", bold=True)
sheet["A1"].font = font_obj1
sheet["A1"] = "Bold Times New Roman"

font_obj2 = Font(size=24, italic=True)
sheet["B3"].font = font_obj2
sheet["B3"] = "24 pt Italic"


# 계산식(Formulas) 적용하기
sheet["C1"] = 200
sheet["C2"] = 300
sheet["C3"] = "=SUM(C1:C2)" # 계산식 설정


# 행의 높이, 열의 너비 설정하기
sheet["A2"] = "Tall row"
sheet["B2"] = "Wide column"
sheet.row_dimensions[2].height = 70
sheet.column_dimensions["B"].width = 20


# 여러개 셀 합치기, 나누기
sheet.merge_cells("A4:D7")
sheet["A4"] = "Twelve cells merged together"
sheet.merge_cells("C8:D8")
sheet["C8"] = "Two merged cells"
sheet.unmerge_cells("C8:D8")


# 고정틀 설정하기
sheet.freeze_panes = "A2"   # 첫 행 고정틀


# 차트 그리기
for i in range(1, 11):
    # 임의 데이터 생성
    sheet[f"E{i}"] = i
ref_obj = openpyxl.chart.Reference(sheet, min_col=5, min_row=1, max_col=5, max_row=10)
series_obj = openpyxl.chart.Series(ref_obj, title="First Series")
chart_obj = openpyxl.chart.BarChart()
chart_obj.title = "My Chart"
chart_obj.append(series_obj)
sheet.add_chart(chart_obj, "G2")

wb.save("sample.xlsx")