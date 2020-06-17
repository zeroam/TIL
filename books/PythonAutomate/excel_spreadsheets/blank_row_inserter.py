"""blank_row_inserter.py
명령행 매개변수로 두 개의 정수와 파일 이름을 문자열로 받는 프로그램
ex) python blank_row_inserter.py 3 2 <*.xlsx>
"""
import sys
import openpyxl

if len(sys.argv) != 4:
    print(f"python {__file__} <start_row> <blank_size> <excel file>")

start_row = int(sys.argv[1])
blank_size = int(sys.argv[2])
file_name = sys.argv[3]

wb = openpyxl.load_workbook(file_name)
sheet = wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for row_obj in sheet.rows:
    for cell_obj in row_obj:
        row_index = cell_obj.row
        col_index = cell_obj.column
        if row_index >= start_row:
            row_index += blank_size
        new_sheet.cell(row=row_index, column=col_index).value = cell_obj.value

new_wb.save(f"blank_{file_name}")
